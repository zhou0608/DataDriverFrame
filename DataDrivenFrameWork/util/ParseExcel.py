#coding=utf-8
import openpyxl
from openpyxl.styles import Border,Side,Font
import time

class ParseExcel(object):
    def __init__(self):
        self.workbook=None
        self.excelFile=None
        self.font=Font(color=None) #设置字体的颜色
        #颜色对应的RGB值
        self.RGBDict={'red':'FFFF3030','green':'FF008B00'}


    def loadWorkBook(self,excelPathAndName):
        #将Excel文件加载到内存，并获取其workbook对象
        try:
            self.workbook=openpyxl.load_workbook(excelPathAndName)
        except Exception as e:
            raise e
        self.excelFile=excelPathAndName
        return self.workbook

    def getSheetByName(self,sheetName):
        #根据sheet名获取该sheet对象
        try:
            sheet=self.workbook.get_sheet_by_name(sheetName)
            return  sheet
        except Exception as e:
            raise e

    def getSheetByIndex(self,sheetIndex):
        #根据sheet的索引号获取该sheet对象
        try:
            sheetname=self.workbook.get_sheet_names()[sheetIndex]
        except Exception as e:
            raise e
        sheet=self.workbook.get_sheet_by_name(sheetname)
        return  sheet

    def getRowNumer(self,sheet):
        #获取sheet中有数据区域的结束行号
        return sheet.max_row

    def getColsNumber(self,sheet):
        #获取sheet中有数据区域的结束列号
        return sheet.max_column

    def getStartRowNumber(self,sheet):
        #获取sheet中有数据区域的开始的行号
        return sheet.min_row

    def getStartColNumber(self,sheet):
        #获取sheet中有数据区域的开始的列号
        return sheet.min_colum

    def getRow(self,sheet,rowNo):
        #获取sheet中某一行，返回的是这一行所有的数据内容组成的tuplue
        #下标从1开始，sheet.rows[1]表示第一行
        try:
            return sheet.rows[rowNo-1]
        except Exception as e:
            raise e

    def getColumn(self,sheet,colNo):
        #获取sheet中某一列，返回的是这一列所有的数据内容组成tuple
        #下标从1开始，sheet.columns[1]表示第一列
        try:
            return sheet.columns[colNo-1]
        except Exception as e:
            raise e
    def getcellOfValue(self,sheet,coordinate=None,rowNo=None,colsNo=None):
        #根据单元格所在的位置索引获取该单元格中的值，下标从1开始
        #sheet.cell(row=1,column=1).value,标识Excel中第一行第一列的值
        if coordinate !=None:
            try:
                return  sheet.cell(coordinate=coordinate).value
            except Exception as e:
                raise  e
        elif coordinate is  None and rowNo is not None and colsNo is not None:
            try:
                return  sheet.cell(row=rowNo,column=colsNo).value
            except Exception as e:
                raise e
        else:
            raise Exception('Insufficient Coordinates of cell!')

    def getCellOfObject(self,sheet,coordinate=None,rowNo=None,colsNo=None):
        #获取某个单元格的对象，可以根据单元格所在位置的数字索引，
        #也可以直接根据Excel中单元格的编码及坐标
        #如getcellobject（sheet,coordinate='A1）or
        #getCellobject(sheet,rowNo=1,colsNo=2)
        if coordinate!=None:
            try:
                return sheet.cell(coordinate=coordinate)
            except Exception as e:
                raise e
        elif coordinate==None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row=rowNo,column=colsNo)
            except Exception as e:
                raise  e
        else:
            raise Exception('Insufficient Coordinates of cell!')

    def writeCell(self,sheet,content,coordinate=None,rowNo=None,colsNo=None,style=None):
        #根据单元格在Excel中的编码坐标或者数字索引坐标向单元格中写入数据
        #下标从1开始，参数style表示字体的颜色的名字，比如red,green
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value = content
                if style is not None:
                    sheet.cell(coordinate=coordinate).font=Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate==None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo,column=colsNo).value=content
                if style:
                    sheet.cell(row=rowNo,column=colsNo).font=Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception('insufficient coordinates of cell!')

    def writeCellCurrentTime(self,sheet,coordinate=None,rowNo=None,colsNo=None):
        #写入当前的时间，下标从1开始
        now=int(time.time())    #显示为时间戳
        timeArray=time.localtime()
        currentTime=time.strftime('%Y -%m -%d %H:%M:%S',timeArray)
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value=currentTime
                sheet.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate==None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo,column=colsNo).value=currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise  e
        else:
            raise  Exception('Insufficient Coordinates of cell!')
if __name__ == '__main__':
    pe= ParseExcel()
    #测试所用的Excel文件‘126邮箱联系人.xlsx’请自行创建
    pe.loadWorkBook(r'D:\File\Python_Code\Book_code\DataDrivenFrameWork\testData\126邮箱联系人.xlsx')
    # print('通过名称获取sheet对象的名字：',pe.getSheetByName(u'联系人').title)
    sheet=pe.getSheetByIndex(0)
    print(type(sheet))
    print('获取最大行号',pe.getRowNumer(sheet))
    print('获取最大列号',pe.getColsNumber(sheet))
    rows=pe.getRow(sheet,1)
    for i in rows:
        print('for value:',i.value)
    pe.writeCell(sheet,u'我爱祖国',rowNo=3,colsNo=2)
    pe.writeCellCurrentTime(sheet,rowNo=3,colsNo=3)